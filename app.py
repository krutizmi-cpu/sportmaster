
from __future__ import annotations

import io
import math
import re
from pathlib import Path
from typing import Dict, Optional, Tuple

import pandas as pd
import streamlit as st
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
DEFAULT_COMMISSIONS_FILENAME = "sportmaster_commissions_2026-02-01.xlsx"

PRODUCT_REQUIRED_COLS = [
    "Артикул",
    "Наименование товара",
    "Себестоимость, ₽",
    "Цена продажи, ₽",
    "Вес факт, кг",
    "Длина, см",
    "Ширина, см",
    "Высота, см",
]

PRODUCT_OPTIONAL_COLS = [
    "Дней хранения FBSM",
    "Реклама, %",
    "Система налогообложения",
    "Налог, %",
    "Прочие расходы, ₽",
    "Целевая маржа, %",
    "Доля возвратов, %",
    "Доля невыкупа/отмен, %",
    "Тарифная группа",
    "Товарная группа 3 уровня",
]

COLUMN_ALIASES = {
    "sku": "Артикул",
    "артикул": "Артикул",
    "наименование": "Наименование товара",
    "наименование товара": "Наименование товара",
    "товар": "Наименование товара",
    "себестоимость": "Себестоимость, ₽",
    "себестоимость руб": "Себестоимость, ₽",
    "цена": "Цена продажи, ₽",
    "цена продажи": "Цена продажи, ₽",
    "price": "Цена продажи, ₽",
    "вес": "Вес факт, кг",
    "вес кг": "Вес факт, кг",
    "вес факт": "Вес факт, кг",
    "длина": "Длина, см",
    "ширина": "Ширина, см",
    "высота": "Высота, см",
    "дней хранения fbsm": "Дней хранения FBSM",
    "реклама": "Реклама, %",
    "система налогообложения": "Система налогообложения",
    "налог": "Налог, %",
    "налог %": "Налог, %",
    "ставка налога %": "Налог, %",
    "ставка налога": "Налог, %",
    "прочие расходы": "Прочие расходы, ₽",
    "целевая маржа": "Целевая маржа, %",
    "доля возвратов": "Доля возвратов, %",
    "доля невыкупа отмен": "Доля невыкупа/отмен, %",
    "тарифная группа": "Тарифная группа",
    "товарная группа 3 уровня": "Товарная группа 3 уровня",
}

TAX_SYSTEMS = {
    "ОСНО (22%)": 22.0,
    "УСН доходы (6%)": 6.0,
    "УСН доходы-расходы (15%)": 15.0,
    "Без налога (0%)": 0.0,
    "Своя ставка": 0.0,
}

STOP_WORDS = {
    "для", "и", "с", "со", "на", "по", "под", "над", "из", "к", "ко", "от", "до", "или", "в", "во",
    "мужские", "мужской", "женские", "женский", "детские", "детский", "взрослые", "взрослый",
    "шт", "комплект", "набор", "унисекс", "товар", "спортивный", "спортивная", "спортивные",
}
TOKEN_SYNONYMS = {
    "велосипеды": "велосипед",
    "велосипедный": "велосипед",
    "велосипедная": "велосипед",
    "велосипедные": "велосипед",
    "велозамок": "замок",
    "велошлем": "шлем",
    "велофонарь": "фонар",
    "велозвонок": "звонок",
    "велонасос": "насос",
    "велодержатель": "держател",
    "вело": "велосипед",
    "кроссовки": "кроссовк",
    "ботинки": "ботинк",
    "перчатки": "перчатк",
    "варежки": "варежк",
    "мячи": "мяч",
    "гантели": "гантел",
    "горный": "горн",
    "беговые": "бег",
    "беговой": "бег",
    "лыжи": "лыж",
    "самокаты": "самокат",
    "замки": "замок",
    "фонари": "фонар",
    "держатели": "держател",
    "крылья": "крыл",
    "насосы": "насос",
    "звонки": "звонок",
    "шлемы": "шлем",
    "седла": "седло",
    "корзины": "корзин",
    "крепления": "креплен",
    "педали": "педал",
}
CATEGORY_OVERRIDE_RULES = [
    {"all": {"велосипед", "горн"}, "lvl3": "Велосипеды"},
    {"all": {"кроссовк", "бег"}, "lvl3": "Кроссовки для бега"},
    {"all": {"велосипед", "замок"}, "lvl3": "Замки для велосипеда"},
    {"all": {"замок"}, "any": {"велосипед"}, "lvl3": "Замки для велосипеда"},
    {"all": {"велосипед", "шлем"}, "lvl3": "Шлемы велосипедные"},
    {"all": {"шлем"}, "any": {"велосипед"}, "lvl3": "Шлемы велосипедные"},
    {"all": {"велосипед", "фонар"}, "lvl3": "Фонари для велосипеда"},
    {"all": {"фонар"}, "any": {"велосипед"}, "lvl3": "Фонари для велосипеда"},
    {"all": {"велосипед", "держател"}, "lvl3": "Держатели для велосипеда"},
    {"all": {"держател"}, "any": {"велосипед"}, "lvl3": "Держатели для велосипеда"},
    {"all": {"велосипед", "звонок"}, "lvl3": "Звонки для велосипеда"},
    {"all": {"велосипед", "крыл"}, "lvl3": "Крылья для велосипеда"},
    {"all": {"велосипед", "корзин"}, "lvl3": "Корзины для велосипеда"},
    {"all": {"велосипед", "насос"}, "lvl3": "Насосы для велосипеда"},
    {"all": {"велосипед", "седло"}, "lvl3": "Седла для велосипеда"},
    {"all": {"велосипед", "креплен"}, "lvl3": "Крепления для велосипеда"},
    {"all": {"велосипед", "педал"}, "lvl3": "Запчасти и компоненты для велосипеда"},
    {"all": {"велосипед", "смазк"}, "lvl3": "Инструменты и смазки для велосипеда"},
]

BASE_COLUMNS = [
    "Артикул", "Наименование товара", "Схема", "Тарифная группа", "Товарная группа 3 уровня", "Как определили категорию",
    "Себестоимость, ₽", "Вес для тарифа, кг", "Налоговый режим", "Ставка налога, %", "Комментарий",
]
CURRENT_PRICE_COLUMNS = [
    "Цена продажи, ₽", "Комиссия, %", "Комиссия, ₽", "Логистика до покупателя, ₽", "Обратная логистика (ожидаемая), ₽",
    "Хранение FBSM, ₽", "Обработка брака, ₽", "Обработка излишков, ₽", "Реклама, %", "Реклама, ₽",
    "Налоговая база, ₽", "Налог, ₽", "Прочие расходы, ₽", "Выплата от МП, ₽", "Полная себестоимость, ₽",
    "Прибыль, ₽", "Маржа к выручке, %", "Наценка на полную себестоимость, %",
]
RECOMMENDED_PRICE_COLUMNS = [
    "Целевая маржа, %", "Рекомендованная цена, ₽", "Комиссия при рекомендованной цене, ₽", "Реклама при рекомендованной цене, ₽",
    "Налог при рекомендованной цене, ₽", "Выплата от МП при рекомендованной цене, ₽",
    "Полная себестоимость при рекомендованной цене, ₽", "Прибыль при рекомендованной цене, ₽", "Маржа при рекомендованной цене, %",
]

def normalize_text(value: str) -> str:
    s = str(value or "").strip().lower().replace("ё", "е")
    s = re.sub(r"[^0-9a-zа-я%]+", " ", s)
    return " ".join(s.split())

def stem_token(token: str) -> str:
    t = normalize_text(token)
    if t in TOKEN_SYNONYMS:
        return TOKEN_SYNONYMS[t]
    for suffix in ["иями", "ями", "ами", "иях", "ия", "ья", "ье", "ий", "ый", "ой", "ая", "яя", "ое", "ее", "ые", "ие", "ам", "ям", "ах", "ях", "ом", "ем", "ую", "юю", "ов", "ев", "ей", "а", "я", "ы", "и", "е", "о", "у", "ю"]:
        if len(t) >= 6 and t.endswith(suffix):
            t = t[:-len(suffix)]
            break
    return TOKEN_SYNONYMS.get(t, t)

def text_tokens(value: str) -> set[str]:
    raw = normalize_text(value).split()
    tokens = {stem_token(tok) for tok in raw if tok and tok not in STOP_WORDS and len(tok) >= 3}
    return {t for t in tokens if t and t not in STOP_WORDS}

def parse_num(value) -> float | None:
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = s.replace("\xa0", "").replace(" ", "")
    s = s.replace("%", "")
    # handle decimals with comma
    if s.count(",") == 1 and s.count(".") == 0:
        s = s.replace(",", ".")
    elif s.count(",") > 1 and s.count(".") == 0:
        s = s.replace(",", "")
    else:
        s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return None

@st.cache_data(show_spinner=False)
def read_reference() -> pd.DataFrame:
    src = DATA_DIR / DEFAULT_COMMISSIONS_FILENAME
    if not src.exists():
        raise FileNotFoundError(f"Не найден файл data/{DEFAULT_COMMISSIONS_FILENAME}. Положите его в repo в папку data.")
    df = pd.read_excel(src)
    df.columns = [str(c).strip() for c in df.columns]
    expected = [
        "Товарная группа 1 уровня", "Товарная группа 2 уровня", "Товарная группа 3 уровня",
        "Тарифная группа", "Ставка комиссии FBSM, %", "Ставка комиссии FBS, %",
    ]
    missing = [c for c in expected if c not in df.columns]
    if missing:
        raise ValueError(f"В файле комиссий отсутствуют колонки: {', '.join(missing)}")
    for col in ["Товарная группа 1 уровня", "Товарная группа 2 уровня", "Товарная группа 3 уровня", "Тарифная группа"]:
        df[col] = df[col].astype(str).str.strip()
        df[f"__norm_{col}"] = df[col].map(normalize_text)
        df[f"__tokens_{col}"] = df[col].map(text_tokens)
    combo_cols = ["Товарная группа 1 уровня", "Товарная группа 2 уровня", "Товарная группа 3 уровня", "Тарифная группа"]
    df["__search_text"] = df[combo_cols].fillna("").astype(str).agg(" ".join, axis=1)
    df["__search_norm"] = df["__search_text"].map(normalize_text)
    df["__tokens_combo"] = df["__search_text"].map(text_tokens)
    for col in ["Ставка комиссии FBSM, %", "Ставка комиссии FBS, %"]:
        df[col] = df[col].map(parse_num).fillna(0.0)
    return df

def prepare_products(df: pd.DataFrame) -> pd.DataFrame:
    renamed = {c: COLUMN_ALIASES.get(normalize_text(c), str(c).strip()) for c in df.columns}
    df = df.rename(columns=renamed).copy()
    missing = [c for c in PRODUCT_REQUIRED_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"В файле товаров не хватает колонок: {', '.join(missing)}")
    for c in PRODUCT_OPTIONAL_COLS:
        if c not in df.columns:
            df[c] = None
    numeric_cols = PRODUCT_REQUIRED_COLS[2:] + ["Дней хранения FBSM", "Реклама, %", "Налог, %", "Прочие расходы, ₽", "Целевая маржа, %", "Доля возвратов, %", "Доля невыкупа/отмен, %"]
    for c in numeric_cols:
        df[c] = df[c].map(parse_num)
    df["Артикул"] = df["Артикул"].fillna("").astype(str).str.strip()
    df["Наименование товара"] = df["Наименование товара"].fillna("").astype(str).str.strip()
    df["Система налогообложения"] = df["Система налогообложения"].fillna("").astype(str).str.strip()

    # remove blank and exported total rows
    mask_blank = (df["Артикул"].eq("") & df["Наименование товара"].eq(""))
    mask_total = df["Артикул"].str.contains("ИТОГО", case=False, na=False)
    df = df.loc[~mask_blank & ~mask_total].copy()
    return df.reset_index(drop=True)

@st.cache_data(show_spinner=False)
def build_product_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    headers = PRODUCT_REQUIRED_COLS + PRODUCT_OPTIONAL_COLS
    ws.append(headers)
    sample_rows = [
        ["SM-001", "Кроссовки беговые мужские", 2500, 5990, 0.8, 32, 22, 12, "", 5, "ОСНО (22%)", "", 0, 20, 0, 0, "", ""],
        ["SM-002", "Велосипед горный", 18000, 32990, 14.5, 145, 25, 78, "", 7, "УСН доходы (6%)", "", 300, 18, 0, 0, "", ""],
        ["SM-003", "Велосипедный замок", 500, 1000, 0.3, 15, 10, 5, "", 0, "ОСНО (22%)", "", 0, 23, 0, 0, "", ""],
    ]
    for row in sample_rows:
        ws.append(row)
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx, col in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(14, min(34, len(col) + 4))
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def normalize_commission_rate(v: float) -> float:
    if pd.isna(v):
        return 0.0
    return float(v * 100.0) if v <= 1 else float(v)

def commission_rate_for_tariff_group(reference_df: pd.DataFrame, tariff_group: str, scheme: str) -> float:
    if not tariff_group:
        return 0.0
    rows = reference_df[reference_df["Тарифная группа"] == tariff_group]
    if rows.empty:
        return 0.0
    col = "Ставка комиссии FBSM, %" if scheme == "FBSM" else "Ставка комиссии FBS, %"
    return normalize_commission_rate(rows.iloc[0][col])

def resolve_tax_rate(row_tax_pct, row_tax_system: str, default_tax_system: str, manual_default_tax_pct: float) -> Tuple[float, str]:
    row_tax_system = str(row_tax_system or "").strip()
    if row_tax_system:
        if row_tax_system == "Своя ставка":
            return float(parse_num(row_tax_pct) or 0.0), row_tax_system
        return float(TAX_SYSTEMS.get(row_tax_system, 0.0)), row_tax_system
    if parse_num(row_tax_pct) is not None and float(parse_num(row_tax_pct)) > 0:
        return float(parse_num(row_tax_pct)), "Своя ставка"
    if default_tax_system == "Своя ставка":
        return float(manual_default_tax_pct), default_tax_system
    return float(TAX_SYSTEMS.get(default_tax_system, 0.0)), default_tax_system

def resolve_row_value(row: pd.Series, col_name: str, sidebar_default: float) -> float:
    value = parse_num(row.get(col_name))
    return float(sidebar_default) if value is None else float(value)

def score_reference_match(name: str, row: pd.Series) -> float:
    name_norm = normalize_text(name)
    name_tokens = text_tokens(name)
    if not name_tokens:
        return 0.0
    best = 0.0
    for col in ["Товарная группа 3 уровня", "Тарифная группа", "Товарная группа 2 уровня"]:
        ref_norm = row[f"__norm_{col}"]
        ref_tokens = row[f"__tokens_{col}"]
        seq = SequenceMatcher(None, name_norm, ref_norm).ratio()
        inter = name_tokens & ref_tokens
        overlap = len(inter) / max(len(ref_tokens), 1)
        cover_name = len(inter) / max(len(name_tokens), 1)
        contains = 1.0 if ref_norm and (ref_norm in name_norm or any(tok in ref_norm for tok in name_tokens if len(tok) >= 5)) else 0.0
        best = max(best, seq * 0.35 + overlap * 0.40 + cover_name * 0.25, overlap * 0.75 + cover_name * 0.25, contains)
    combo_tokens = row["__tokens_combo"]
    combo_norm = row["__search_norm"]
    inter = name_tokens & combo_tokens
    combo_overlap = len(inter) / max(len(name_tokens), 1)
    combo_cover = len(inter) / max(len(combo_tokens), 1)
    prefix_hits = sum(1 for tok in name_tokens for ref_tok in combo_tokens if tok == ref_tok or tok.startswith(ref_tok) or ref_tok.startswith(tok))
    prefix_boost = min(prefix_hits * 0.04, 0.20)
    phrase_boost = 0.0
    if {"велосипед", "замок"}.issubset(name_tokens) and "замки для велосипеда" in combo_norm:
        phrase_boost += 0.45
    if {"велосипед", "горн"}.issubset(name_tokens) and "велосипеды" in combo_norm:
        phrase_boost += 0.35
    if {"кроссовк", "бег"}.issubset(name_tokens) and "кроссовки для бега" in combo_norm:
        phrase_boost += 0.35
    if "велосипед" in name_tokens and "аксессуары для велоспорта" in combo_norm:
        phrase_boost += 0.12
    if inter and len(inter) >= 2:
        phrase_boost += 0.10
    best = max(best, combo_overlap * 0.70 + combo_cover * 0.12 + prefix_boost + phrase_boost)
    return min(best, 0.99)

def resolve_override_rule(reference_df: pd.DataFrame, product_name: str) -> Tuple[str, str, Optional[float], str] | None:
    name_tokens = text_tokens(product_name)
    if not name_tokens:
        return None
    for rule in CATEGORY_OVERRIDE_RULES:
        all_req = set(rule.get("all", set()))
        any_req = set(rule.get("any", set()))
        if all_req and not all_req.issubset(name_tokens):
            continue
        if any_req and not (any_req & name_tokens):
            continue
        rows = reference_df[reference_df["Товарная группа 3 уровня"] == rule["lvl3"]]
        if not rows.empty:
            row = rows.iloc[0]
            return str(row["Тарифная группа"]), str(row["Товарная группа 3 уровня"]), 0.99, "Правило по ключевым словам"
    return None

def resolve_tariff_group(reference_df: pd.DataFrame, product_name: str, manual_tariff_group: str, manual_group3: str) -> Tuple[str, str, Optional[float], str]:
    name_norm = normalize_text(product_name)
    manual_tariff_group = str(manual_tariff_group or "").strip()
    manual_group3 = str(manual_group3 or "").strip()
    if manual_tariff_group:
        rows = reference_df[reference_df["Тарифная группа"] == manual_tariff_group]
        if not rows.empty:
            group3 = str(rows.iloc[0]["Товарная группа 3 уровня"])
            return manual_tariff_group, group3, 1.0, "Тарифная группа из файла товаров"
    if manual_group3:
        rows = reference_df[reference_df["Товарная группа 3 уровня"] == manual_group3]
        if not rows.empty:
            tariff_group = str(rows.iloc[0]["Тарифная группа"])
            return tariff_group, manual_group3, 1.0, "Товарная группа 3 уровня из файла товаров"
    override = resolve_override_rule(reference_df, product_name)
    if override is not None:
        return override
    for col, note in [
        ("Товарная группа 3 уровня", "Точное совпадение с ТГ3"),
        ("Тарифная группа", "Точное совпадение с тарифной группой"),
        ("Товарная группа 2 уровня", "Точное совпадение со 2 уровнем"),
    ]:
        rows = reference_df[reference_df[f"__norm_{col}"] == name_norm]
        if not rows.empty:
            row = rows.iloc[0]
            return str(row["Тарифная группа"]), str(row["Товарная группа 3 уровня"]), 1.0, note
    candidates = []
    for _, ref_row in reference_df.iterrows():
        score = score_reference_match(product_name, ref_row)
        if score > 0:
            candidates.append((score, str(ref_row["Тарифная группа"]), str(ref_row["Товарная группа 3 уровня"])))
    if not candidates:
        return "", "", None, "Не найдено"
    candidates.sort(key=lambda x: x[0], reverse=True)
    best_score, best_tariff, best_lvl3 = candidates[0]
    if best_score >= 0.42:
        return best_tariff, best_lvl3, best_score, f"Автоподбор по названию ({best_score:.2f})"
    return "", "", best_score, "Не найдено"

def ceil_to(value: float, step: float) -> float:
    return 0.0 if value <= 0 else math.ceil(value / step) * step

def calc_fbs_billable_weight(weight_kg: float, length_cm: float, width_cm: float, height_cm: float, basis: str) -> float:
    volume = max(length_cm, 0) * max(width_cm, 0) * max(height_cm, 0)
    actual = max(weight_kg, 0)
    pr = max(actual, volume / 5000.0)
    cdek = max(actual, volume / 4000.0)
    chosen = {"Почта России": pr, "СДЭК": cdek, "Фактический": actual}.get(basis, max(pr, cdek))
    return float(math.ceil(chosen))

def calc_fbs_delivery(weight_kg: float, profile: str) -> float:
    if weight_kg <= 0:
        return 0.0
    base, extra = (220.0, 90.0) if profile == "220 + 90" else (200.0, 70.0)
    return base if weight_kg <= 2 else base + (weight_kg - 2) * extra

def calc_fbsm_delivery(weight_kg: float, profile: str) -> float:
    w = ceil_to(weight_kg, 0.1)
    if w <= 0:
        return 0.0
    up05, up10, extra = (56.0, 90.0, 60.0) if profile == "56 / 90 / 60" else (35.0, 75.0, 35.0)
    if w <= 0.5:
        return up05
    if w <= 1.0:
        return up10
    return up10 + (w - 1.0) * extra

def calc_fbsm_storage(weight_kg: float, storage_days: float) -> float:
    days = max(int(storage_days or 0), 0)
    if days <= 60:
        return 0.0
    if days <= 90:
        return max(weight_kg, 0) * (days - 60) * 3.0
    return max(weight_kg, 0) * 30 * 3.0 + max(weight_kg, 0) * (days - 90) * 6.0

def tax_mode(tax_system_label: str) -> str:
    if tax_system_label == "ОСНО (22%)":
        return "profit"
    if tax_system_label == "УСН доходы-расходы (15%)":
        return "profit"
    return "revenue"

def calc_tax_amount(price: float, tax_pct: float, tax_system_label: str, profit_before_tax: float) -> Tuple[float, float]:
    rate = tax_pct / 100.0
    if rate <= 0 or price <= 0:
        return 0.0, 0.0
    mode = tax_mode(tax_system_label)
    if mode == "profit":
        base = max(profit_before_tax, 0.0)
    else:
        base = max(price, 0.0)
    return base, base * rate

def solve_target_price(target_margin_pct: float, fixed_cost_rub: float, commission_pct: float, ads_pct: float, tax_pct: float, tax_system_label: str) -> Optional[float]:
    t = target_margin_pct / 100.0
    c = commission_pct / 100.0
    a = ads_pct / 100.0
    r = tax_pct / 100.0
    mode = tax_mode(tax_system_label)
    if mode == "profit":
        denominator = (1 - r) * (1 - c - a) - t
        return None if denominator <= 0 else (1 - r) * fixed_cost_rub / denominator
    denominator = 1 - c - a - r - t
    return None if denominator <= 0 else fixed_cost_rub / denominator

def calculate_row(
    row: pd.Series,
    reference_df: pd.DataFrame,
    scheme: str,
    fbs_weight_basis: str,
    fbs_logistics_profile: str,
    fbsm_logistics_profile: str,
    include_fbsm_return_logistics: bool,
    include_fbsm_defect_handling: bool,
    include_fbsm_excess_handling: bool,
    default_tax_system: str,
    manual_default_tax_pct: float,
    sidebar_defaults: dict,
) -> Dict:
    sku = str(row["Артикул"] or "").strip()
    name = str(row["Наименование товара"] or "").strip()
    cost = float(parse_num(row["Себестоимость, ₽"]) or 0.0)
    price = float(parse_num(row["Цена продажи, ₽"]) or 0.0)
    actual_weight = float(parse_num(row["Вес факт, кг"]) or 0.0)
    length = float(parse_num(row["Длина, см"]) or 0.0)
    width = float(parse_num(row["Ширина, см"]) or 0.0)
    height = float(parse_num(row["Высота, см"]) or 0.0)
    storage_days = resolve_row_value(row, "Дней хранения FBSM", sidebar_defaults["Дней хранения FBSM"])
    ads_pct = resolve_row_value(row, "Реклама, %", sidebar_defaults["Реклама, %"])
    other_costs = resolve_row_value(row, "Прочие расходы, ₽", sidebar_defaults["Прочие расходы, ₽"])
    target_margin = resolve_row_value(row, "Целевая маржа, %", sidebar_defaults["Целевая маржа, %"])
    return_rate = resolve_row_value(row, "Доля возвратов, %", sidebar_defaults["Доля возвратов, %"])
    cancel_rate = resolve_row_value(row, "Доля невыкупа/отмен, %", sidebar_defaults["Доля невыкупа/отмен, %"])
    row_tax_pct = row.get("Налог, %")
    row_tax_system = str(row.get("Система налогообложения", "") or "")

    tariff_group, lvl3_group, match_score, match_note = resolve_tariff_group(reference_df, name, row.get("Тарифная группа", ""), row.get("Товарная группа 3 уровня", ""))
    commission_pct = commission_rate_for_tariff_group(reference_df, tariff_group, scheme) or 0.0
    tax_pct, tax_system_label = resolve_tax_rate(row_tax_pct, row_tax_system, default_tax_system, manual_default_tax_pct)

    commission_rub = price * commission_pct / 100.0
    ad_cost_rub = price * ads_pct / 100.0
    logistics_to_buyer = reverse_logistics = storage_rub = defect_handling = excess_handling = billable_weight = 0.0

    if scheme == "FBS":
        billable_weight = calc_fbs_billable_weight(actual_weight, length, width, height, fbs_weight_basis)
        logistics_to_buyer = calc_fbs_delivery(billable_weight, fbs_logistics_profile)
    else:
        billable_weight = ceil_to(actual_weight, 0.1)
        logistics_to_buyer = calc_fbsm_delivery(billable_weight, fbsm_logistics_profile)
        if include_fbsm_return_logistics:
            reverse_logistics = logistics_to_buyer * (return_rate + cancel_rate) / 100.0
        storage_rub = calc_fbsm_storage(actual_weight, storage_days)
        defect_handling = 30.0 if include_fbsm_defect_handling else 0.0
        excess_handling = 30.0 if include_fbsm_excess_handling else 0.0

    mp_services_rub = logistics_to_buyer + reverse_logistics + storage_rub + defect_handling + excess_handling
    payout_from_mp = price - commission_rub - mp_services_rub
    profit_before_tax = price - commission_rub - mp_services_rub - ad_cost_rub - other_costs - cost
    tax_base_rub, tax_rub = calc_tax_amount(price, tax_pct, tax_system_label, profit_before_tax)
    full_cost = cost + commission_rub + mp_services_rub + ad_cost_rub + tax_rub + other_costs
    profit = price - full_cost
    margin_on_revenue_pct = (profit / price * 100.0) if price else 0.0
    markup_on_full_cost_pct = ((price / full_cost) - 1.0) * 100.0 if full_cost else 0.0

    target_price = target_profit_rub = target_margin_pct_fact = None
    target_commission = target_ads = target_tax = target_payout_from_mp = target_full_cost = None
    fixed_costs = cost + mp_services_rub + other_costs
    if target_margin > 0:
        target_price = solve_target_price(target_margin, fixed_costs, commission_pct, ads_pct, tax_pct, tax_system_label)
        if target_price is not None:
            target_commission = target_price * commission_pct / 100.0
            target_ads = target_price * ads_pct / 100.0
            target_profit_before_tax = target_price - target_commission - mp_services_rub - target_ads - other_costs - cost
            _, target_tax = calc_tax_amount(target_price, tax_pct, tax_system_label, target_profit_before_tax)
            target_payout_from_mp = target_price - target_commission - mp_services_rub
            target_full_cost = cost + mp_services_rub + other_costs + target_commission + target_ads + target_tax
            target_profit_rub = target_price - target_full_cost
            target_margin_pct_fact = (target_profit_rub / target_price * 100.0) if target_price else None

    warnings = []
    if not tariff_group:
        warnings.append("Не определена тарифная группа")
    if commission_pct == 0:
        warnings.append("Комиссия не найдена")
    if target_margin <= 0:
        warnings.append("Целевая маржа не задана")
    elif target_price is None:
        warnings.append("Рекомендованная цена не считается: слишком высокая сумма комиссии, рекламы, налога и целевой маржи")
    if price <= 0:
        warnings.append("Цена продажи 0")
    if actual_weight <= 0:
        warnings.append("Вес товара 0")

    return {
        "Артикул": sku,
        "Наименование товара": name,
        "Схема": scheme,
        "Тарифная группа": tariff_group,
        "Товарная группа 3 уровня": lvl3_group,
        "Как определили категорию": match_note,
        "Себестоимость, ₽": round(cost, 2),
        "Вес для тарифа, кг": round(billable_weight, 2),
        "Налоговый режим": tax_system_label,
        "Ставка налога, %": round(tax_pct, 2),
        "Цена продажи, ₽": round(price, 2),
        "Комиссия, %": round(commission_pct, 2),
        "Комиссия, ₽": round(commission_rub, 2),
        "Логистика до покупателя, ₽": round(logistics_to_buyer, 2),
        "Обратная логистика (ожидаемая), ₽": round(reverse_logistics, 2),
        "Хранение FBSM, ₽": round(storage_rub, 2),
        "Обработка брака, ₽": round(defect_handling, 2),
        "Обработка излишков, ₽": round(excess_handling, 2),
        "Реклама, %": round(ads_pct, 2),
        "Реклама, ₽": round(ad_cost_rub, 2),
        "Налоговая база, ₽": round(tax_base_rub, 2),
        "Налог, ₽": round(tax_rub, 2),
        "Прочие расходы, ₽": round(other_costs, 2),
        "Выплата от МП, ₽": round(payout_from_mp, 2),
        "Полная себестоимость, ₽": round(full_cost, 2),
        "Прибыль, ₽": round(profit, 2),
        "Маржа к выручке, %": round(margin_on_revenue_pct, 2),
        "Наценка на полную себестоимость, %": round(markup_on_full_cost_pct, 2),
        "Целевая маржа, %": round(target_margin, 2),
        "Рекомендованная цена, ₽": round(target_price, 2) if target_price is not None else None,
        "Комиссия при рекомендованной цене, ₽": round(target_commission, 2) if target_commission is not None else None,
        "Реклама при рекомендованной цене, ₽": round(target_ads, 2) if target_ads is not None else None,
        "Налог при рекомендованной цене, ₽": round(target_tax, 2) if target_tax is not None else None,
        "Выплата от МП при рекомендованной цене, ₽": round(target_payout_from_mp, 2) if target_payout_from_mp is not None else None,
        "Полная себестоимость при рекомендованной цене, ₽": round(target_full_cost, 2) if target_full_cost is not None else None,
        "Прибыль при рекомендованной цене, ₽": round(target_profit_rub, 2) if target_profit_rub is not None else None,
        "Маржа при рекомендованной цене, %": round(target_margin_pct_fact, 2) if target_margin_pct_fact is not None else None,
        "Комментарий": "; ".join(warnings),
    }

def grouped_display_df(result_df: pd.DataFrame) -> pd.DataFrame:
    columns = []
    for col in BASE_COLUMNS:
        if col in result_df.columns:
            columns.append(("Карточка товара", col))
    for col in CURRENT_PRICE_COLUMNS:
        if col in result_df.columns:
            columns.append(("Расчет по текущей цене", col))
    for col in RECOMMENDED_PRICE_COLUMNS:
        if col in result_df.columns:
            columns.append(("Расчет по рекомендованной цене", col))
    disp = result_df[[c[1] for c in columns]].copy()
    disp.columns = pd.MultiIndex.from_tuples(columns)
    return disp

def build_export_workbook(result_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"

    grouped = [("Карточка товара", BASE_COLUMNS), ("Расчет по текущей цене", CURRENT_PRICE_COLUMNS), ("Расчет по рекомендованной цене", RECOMMENDED_PRICE_COLUMNS)]
    flat_cols = [c for _, cols in grouped for c in cols if c in result_df.columns]
    row1, row2 = [], []
    merge_ranges = []
    current_col = 1
    for title, cols in grouped:
        use_cols = [c for c in cols if c in result_df.columns]
        if not use_cols:
            continue
        start_col = current_col
        for col in use_cols:
            row1.append(title)
            row2.append(col)
            current_col += 1
        end_col = current_col - 1
        if end_col > start_col:
            merge_ranges.append((start_col, end_col, title))
    ws.append(row1)
    ws.append(row2)
    for row in result_df[flat_cols].itertuples(index=False):
        ws.append(list(row))

    group_fills = {
        "Карточка товара": PatternFill("solid", fgColor="1F4E78"),
        "Расчет по текущей цене": PatternFill("solid", fgColor="5B9BD5"),
        "Расчет по рекомендованной цене": PatternFill("solid", fgColor="70AD47"),
    }
    second_fill = PatternFill("solid", fgColor="D9EAF7")
    for c in range(1, len(flat_cols) + 1):
        group_name = ws.cell(1, c).value
        ws.cell(1, c).fill = group_fills.get(group_name, group_fills["Карточка товара"])
        ws.cell(1, c).font = Font(color="FFFFFF", bold=True)
        ws.cell(2, c).fill = second_fill
        ws.cell(2, c).font = Font(bold=True)
        ws.cell(1, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(2, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for start_col, end_col, _ in merge_ranges:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    money_cols = {c for c in flat_cols if "₽" in c}
    pct_cols = {
        "Комиссия, %", "Реклама, %", "Ставка налога, %", "Маржа к выручке, %",
        "Наценка на полную себестоимость, %", "Целевая маржа, %", "Маржа при рекомендованной цене, %",
    }
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            col_name = ws.cell(2, cell.column).value
            if col_name in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00_);[Red](#,##0.00)'
            elif col_name in pct_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    total_row = ws.max_row + 2
    ws.cell(total_row, 1).value = "ИТОГО / СРЕДНЕЕ"
    total_fill = PatternFill("solid", fgColor="E2F0D9")
    for idx, col_name in enumerate(flat_cols, start=1):
        col_letter = get_column_letter(idx)
        cell = ws.cell(total_row, idx)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        if col_name in money_cols:
            cell.value = f"=SUM({col_letter}3:{col_letter}{total_row - 2})"
            cell.number_format = '#,##0.00_);[Red](#,##0.00)'
        elif col_name in pct_cols:
            cell.value = f"=AVERAGE({col_letter}3:{col_letter}{total_row - 2})"
            cell.number_format = '0.00'
    for col_idx, col_name in enumerate(flat_cols, start=1):
        max_len = max(len(str(col_name)), len(str(ws.cell(1, col_idx).value or "")))
        for row_idx in range(2, ws.max_row + 1):
            max_len = max(max_len, len(str(ws.cell(row_idx, col_idx).value or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 34)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(flat_cols))}{ws.max_row}"

    guide = wb.create_sheet("Пояснения")
    guide.append(["Показатель", "Описание"])
    guide_rows = [
        ("Выплата от МП, ₽", "Сумма после комиссии и услуг маркетплейса. Себестоимость товара, реклама, налог и прочие расходы в эту колонку не входят."),
        ("Полная себестоимость, ₽", "Себестоимость товара + комиссия + логистика/хранение/обработка + реклама + налог + прочие расходы."),
        ("Маржа к выручке, %", "Прибыль / цена продажи * 100."),
        ("Наценка на полную себестоимость, %", "(Цена продажи / полная себестоимость - 1) * 100."),
        ("Рекомендованная цена, ₽", "Цена, которая нужна для достижения целевой маржи с учетом комиссии, логистики, рекламы, налога и прочих расходов."),
    ]
    for row in guide_rows:
        guide.append(row)
    for col_idx in range(1, 3):
        guide.column_dimensions[get_column_letter(col_idx)].width = 48 if col_idx == 2 else 34

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def render_overview_metrics(result_df: pd.DataFrame) -> None:
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("SKU", int(len(result_df)))
    with c2:
        st.metric("Суммарная прибыль", f"{result_df['Прибыль, ₽'].fillna(0).sum():,.2f} ₽")
    with c3:
        avg_margin = result_df["Маржа к выручке, %"].fillna(0).mean() if len(result_df) else 0.0
        st.metric("Средняя маржа", f"{avg_margin:.1f}%")
    with c4:
        avg_target = result_df["Рекомендованная цена, ₽"].dropna().mean()
        st.metric("Средняя рекомендованная цена", "—" if pd.isna(avg_target) else f"{avg_target:,.2f} ₽")

def app() -> None:
    st.set_page_config(page_title="Спортмастер — юнит-экономика", page_icon="🏃", layout="wide")
    st.title("🏃 Спортмастер — юнит-экономика")
    st.caption("Массовая загрузка товаров через Excel • автоматическое определение категории • моментальный пересчет комиссии, налога, выплаты от МП и рекомендованной цены")

    try:
        reference_df = read_reference()
    except Exception as e:
        st.error(str(e))
        st.stop()

    with st.sidebar:
        st.header("Параметры расчета")
        scheme = st.selectbox("Схема", ["FBS", "FBSM"], index=0)

        st.subheader("Налогообложение")
        default_tax_system = st.selectbox("Система налогообложения", list(TAX_SYSTEMS.keys()), index=0)
        manual_default_tax_pct = 0.0
        if default_tax_system == "Своя ставка":
            manual_default_tax_pct = st.number_input("Своя ставка налога, %", min_value=0.0, value=6.0, step=0.5)

        st.subheader("Параметры по умолчанию")
        default_target_margin_pct = st.number_input("Целевая маржа по умолчанию, %", min_value=0.0, value=20.0, step=1.0)
        default_ads_pct = st.number_input("Реклама по умолчанию, %", min_value=0.0, value=0.0, step=0.5)
        default_storage_days = st.number_input("Дней хранения FBSM по умолчанию", min_value=0.0, value=0.0, step=1.0)
        default_other_costs = st.number_input("Прочие расходы по умолчанию, ₽", min_value=0.0, value=0.0, step=50.0)
        default_returns_pct = st.number_input("Доля возвратов по умолчанию, %", min_value=0.0, value=0.0, step=0.5)
        default_cancel_pct = st.number_input("Доля невыкупа/отмен по умолчанию, %", min_value=0.0, value=0.0, step=0.5)

        st.subheader("Логистика FBS")
        fbs_weight_basis_ui = st.selectbox("Основа веса", ["Консервативно (макс из Почты/СДЭК)", "Почта России", "СДЭК", "Фактический"], index=0)
        fbs_weight_basis_map = {
            "Консервативно (макс из Почты/СДЭК)": "Консервативно",
            "Почта России": "Почта России",
            "СДЭК": "СДЭК",
            "Фактический": "Фактический",
        }
        fbs_logistics_profile = st.selectbox("Тариф FBS", ["200 + 70", "220 + 90"], index=0)

        st.subheader("Параметры FBSM")
        fbsm_logistics_profile = st.selectbox("Тариф FBSM", ["35 / 75 / 35", "56 / 90 / 60"], index=0)
        include_fbsm_return_logistics = st.checkbox("Учитывать обратную логистику FBSM", value=True)
        include_fbsm_defect_handling = st.checkbox("Учитывать обработку брака 30 ₽/шт", value=False)
        include_fbsm_excess_handling = st.checkbox("Учитывать обработку излишков 30 ₽/шт", value=False)

    sidebar_defaults = {
        "Дней хранения FBSM": default_storage_days,
        "Реклама, %": default_ads_pct,
        "Прочие расходы, ₽": default_other_costs,
        "Целевая маржа, %": default_target_margin_pct,
        "Доля возвратов, %": default_returns_pct,
        "Доля невыкупа/отмен, %": default_cancel_pct,
    }

    st.download_button(
        "Скачать шаблон товаров Excel",
        data=build_product_template_bytes(),
        file_name="Шаблон_товаров_Спортмастер.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )
    products_file = st.file_uploader("Загрузите Excel-файл с товарами", type=["xlsx"])
    st.caption("Цена влияет на комиссию, рекламу, налог, выплату от МП, прибыль и рекомендованную цену. Логистика зависит от веса и габаритов, поэтому от цены не меняется.")

    if products_file is None:
        st.info("Сначала скачайте шаблон, заполните товары и загрузите файл.")
        return

    try:
        products_df = prepare_products(pd.read_excel(products_file))
    except Exception as e:
        st.error(f"Ошибка в файле товаров: {e}")
        st.stop()

    if products_df.empty:
        st.warning("В файле не найдено строк с товарами.")
        return

    result_rows = []
    for _, row in products_df.iterrows():
        result_rows.append(calculate_row(
            row, reference_df, scheme, fbs_weight_basis_map[fbs_weight_basis_ui], fbs_logistics_profile,
            fbsm_logistics_profile, include_fbsm_return_logistics, include_fbsm_defect_handling,
            include_fbsm_excess_handling, default_tax_system, manual_default_tax_pct, sidebar_defaults
        ))
    result_df = pd.DataFrame(result_rows)
    render_overview_metrics(result_df)

    disp = grouped_display_df(result_df)
    st.dataframe(disp, use_container_width=True, hide_index=True, height=700)

    export_bytes = build_export_workbook(result_df)
    st.download_button(
        "Скачать результат в Excel",
        data=export_bytes,
        file_name=f"Спортмастер_юнит-экономика_{scheme}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

if __name__ == "__main__":
    app()
